import os
import logging
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from supabase import create_client, Client
from dotenv import load_dotenv

# 1. Configuración
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

load_dotenv()
SUPABASE_URL = os.environ.get("SUPABASE_URL")
SUPABASE_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
TENANT_ID = os.environ.get("TENANT_ID")

if not SUPABASE_URL or not SUPABASE_KEY:
    logger.error("Faltan variables de entorno.")
    exit(1)

supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

def clean_column_names(columns: list) -> list:
    # Aseguramos que los nombres sean consistentes y sin caracteres especiales complejos
    return pd.Series(columns).astype(str).str.strip().str.lower().str.replace(' ', '_').str.replace('ñ', 'n').str.replace('ü', 'u').str.replace(r'[^\w\s]', '', regex=True).tolist()

# --- NUEVO: FUNCIÓN DE LIMPIEZA DE TEXTOS ---
def limpiar_texto(texto):
    """Limpia nombres de materias, quitando dobles espacios y arreglando errores ortográficos como CALUCLO."""
    if pd.isna(texto):
        return "DESCONOCIDO"
    texto = str(texto).upper().strip()
    texto = ' '.join(texto.split()) # Quita dobles espacios
    texto = texto.replace("CALUCLO", "CÁLCULO") # Parche para el error de Cálculo Diferencial
    return texto

def upsert_batch(table_name: str, records: list, conflict_columns: str):
    if not records: return
    
    # Sanitización centralizada: Reemplazar NaN por None y convertir tipos de Numpy/Pandas
    cleaned_records = []
    for record in records:
        clean_record = {}
        for k, v in record.items():
            if pd.isna(v):
                clean_record[k] = None
            elif hasattr(v, 'item'): # Extrae el tipo nativo de Python (ej: Int64 -> int)
                val = v.item()
                clean_record[k] = val if pd.notna(val) else None
            else:
                clean_record[k] = v
        cleaned_records.append(clean_record)

    try:
        supabase.table(table_name).upsert(cleaned_records, on_conflict=conflict_columns).execute()
    except Exception as e:
        logger.error(f"Error en tabla {table_name}: {e}")
        raise e

def process_and_upload(df_chunk: pd.DataFrame):
    df = df_chunk.copy()
    if TENANT_ID: df['tenant_id'] = TENANT_ID
    
    # Normalización de textos y códigos
    for col in ['documento', 'grupo', 'codigo_asignatura']:
        if col in df.columns:
            df[col] = df[col].astype(str).str.replace('.0', '', regex=False).str.strip()

    # --- NUEVO: APLICAR LIMPIEZA A LAS ASIGNATURAS ---
    if 'asignatura' in df.columns:
        df['asignatura'] = df['asignatura'].apply(limpiar_texto)

    if 'nombres_docente' in df.columns:
        df['docente_full_name'] = (df['nombres_docente'].fillna('') + ' ' + df['apellidos_docente'].fillna('')).str.strip()

    # 1. CAMPUSES (Sedes) - Unión de Sede Estudiante y Sede de clase
    campus_list = []
    if 'sede_estudiante' in df.columns: campus_list.extend(df['sede_estudiante'].dropna().unique().tolist())
    if 'sede' in df.columns: campus_list.extend(df['sede'].dropna().unique().tolist())
    
    if campus_list:
        campuses_df = pd.DataFrame({'name': list(set(campus_list)), 'tenant_id': TENANT_ID})
        upsert_batch('campuses', campuses_df.to_dict('records'), 'tenant_id, name')
    
    res_campuses = supabase.table('campuses').select('id, name').eq('tenant_id', TENANT_ID).execute()
    map_campuses = {r['name']: r['id'] for r in res_campuses.data}

    # 2. FACULTIES (Facultades)
    if 'facultad_grupo' in df.columns:
        faculties_df = df[['facultad_grupo', 'tenant_id']].rename(columns={'facultad_grupo': 'name'}).drop_duplicates().dropna()
        upsert_batch('faculties', faculties_df.to_dict('records'), 'tenant_id, name')
    
    res_faculties = supabase.table('faculties').select('id, name').eq('tenant_id', TENANT_ID).execute()
    map_faculties = {r['name']: r['id'] for r in res_faculties.data}

    # 3. ACADEMIC PROGRAMS
    if 'programa' in df.columns:
        programs_df = df[['programa', 'facultad_grupo', 'tenant_id']].copy()
        programs_df = programs_df.rename(columns={'programa': 'name'})
        programs_df['faculty_id'] = programs_df['facultad_grupo'].map(map_faculties)
        programs_df = programs_df[['name', 'faculty_id', 'tenant_id']].drop_duplicates().dropna(subset=['name', 'faculty_id'])
        upsert_batch('academic_programs', programs_df.to_dict('records'), 'tenant_id, name')

    res_programs = supabase.table('academic_programs').select('id, name').eq('tenant_id', TENANT_ID).execute()
    map_programs = {r['name']: r['id'] for r in res_programs.data}

    # 4. TEACHERS & SUBJECTS
    if 'docente_full_name' in df.columns:
        teachers_df = df[['docente_full_name', 'tenant_id']].copy().rename(columns={'docente_full_name': 'full_name'})
        if 'departamento' in df.columns:
            teachers_df['department'] = df['departamento']
        teachers_df = teachers_df.drop_duplicates('full_name').dropna(subset=['full_name'])
        teachers_df = teachers_df[teachers_df['full_name'] != '']
        upsert_batch('teachers', teachers_df.to_dict('records'), 'tenant_id, full_name')

    if 'codigo_asignatura' in df.columns:
        subjects_df = df[['codigo_asignatura', 'asignatura', 'creditos', 'tenant_id']].copy()
        subjects_df = subjects_df.rename(columns={'codigo_asignatura': 'code', 'asignatura': 'name', 'creditos': 'credits'})
        subjects_df['credits'] = pd.to_numeric(subjects_df['credits'], errors='coerce').fillna(0).astype(int)
        subjects_df = subjects_df.drop_duplicates('code').dropna(subset=['code'])
        upsert_batch('subjects', subjects_df.to_dict('records'), 'tenant_id, code')

    res_teach = supabase.table('teachers').select('id, full_name').eq('tenant_id', TENANT_ID).execute()
    map_teachers = {r['full_name']: r['id'] for r in res_teach.data}

    res_subj = supabase.table('subjects').select('id, code').eq('tenant_id', TENANT_ID).execute()
    map_subjects = {r['code']: r['id'] for r in res_subj.data}

    # 5. STUDENTS (Enriquecido)
    if 'documento' in df.columns:
        students_df = df[['documento', 'nombres', 'apellidos', 'correo_electronico', 'sexo', 'estrato', 'barrio', 'comuna', 'direccion', 'antiguedad', 'programa', 'sede_estudiante', 'tenant_id']].copy()
        students_df['full_name'] = (students_df['nombres'].fillna('') + ' ' + students_df['apellidos'].fillna('')).str.strip()
        
        students_df = students_df.rename(columns={
            'documento': 'external_id', 
            'correo_electronico': 'email',
            'sexo': 'gender',
            'estrato': 'stratum',
            'direccion': 'address'
        })
        
        # Limpieza de numéricos y texto categórico
        if 'stratum' in students_df.columns:
            students_df['stratum'] = pd.to_numeric(students_df['stratum'], errors='coerce').astype('Int64')
        
        # Antigüedad como variable categórica
        if 'antiguedad' in students_df.columns:
            students_df['antiguedad'] = students_df['antiguedad'].astype(str).str.strip()
            students_df['antiguedad'] = students_df['antiguedad'].replace(['nan', 'NaN', 'None', '<NA>', ''], None)
        
        # Mapeo de FKs
        students_df['program_id'] = students_df['programa'].map(map_programs)
        students_df['campus_id'] = students_df['sede_estudiante'].map(map_campuses)
        
        final_students = students_df[['external_id', 'full_name', 'email', 'gender', 'stratum', 'barrio', 'comuna', 'address', 'antiguedad', 'program_id', 'campus_id', 'tenant_id']]
        final_students = final_students.drop_duplicates('external_id').dropna(subset=['external_id'])
        upsert_batch('students', final_students.to_dict('records'), 'tenant_id, external_id')

    res_stud = supabase.table('students').select('id, external_id').eq('tenant_id', TENANT_ID).execute()
    map_students = {r['external_id']: r['id'] for r in res_stud.data}

    # 6. CLASS GROUPS
    if 'grupo' in df.columns:
        groups_df = df[['grupo', 'ano', 'semestre', 'codigo_asignatura', 'docente_full_name', 'tenant_id']].copy()
        groups_df['semester'] = groups_df['ano'].astype(str).str.replace('.0','') + '-' + groups_df['semestre'].astype(str).str.replace('.0','')
        groups_df['semester'] = groups_df['semester'].replace(['nan-nan', 'nan-1', 'nan-2', '2024-nan', '2025-nan'], None)
        
        groups_df = groups_df.rename(columns={'grupo': 'group_code'})
        groups_df['subject_id'] = groups_df['codigo_asignatura'].map(map_subjects)
        groups_df['teacher_id'] = groups_df['docente_full_name'].map(map_teachers)
        
        if 'modalidad' in df.columns:
            modalidad_map = {
                'presencial': 'Presencial',
                'teams': 'Teams',
                'virtual': 'Teams',
                'híbrido': 'Híbrido',
                'hibrido': 'Híbrido'
            }
            groups_df['modality'] = df['modalidad'].astype(str).str.strip().str.lower().map(modalidad_map)
            groups_df['modality'] = groups_df['modality'].replace({pd.NA: None, np.nan: None})
            
        groups_df = groups_df[['group_code', 'semester', 'subject_id', 'teacher_id', 'tenant_id', 'modality'] if 'modalidad' in df.columns else ['group_code', 'semester', 'subject_id', 'teacher_id', 'tenant_id']]
        groups_df = groups_df.dropna(subset=['group_code', 'subject_id', 'semester']).drop_duplicates(['group_code', 'subject_id', 'semester'])
        upsert_batch('class_groups', groups_df.to_dict('records'), 'tenant_id, subject_id, semester, group_code')

    res_groups = supabase.table('class_groups').select('id, group_code, subject_id, semester').eq('tenant_id', TENANT_ID).execute()
    map_groups = {(r['group_code'], r['subject_id'], r['semester']): r['id'] for r in res_groups.data}

    # 7. GROUP SCHEDULES
    if 'dia' in df.columns:
        sched_df = df.copy()
        sched_df['semester'] = sched_df['ano'].astype(str).str.replace('.0','') + '-' + sched_df['semestre'].astype(str).str.replace('.0','')
        sched_df['semester'] = sched_df['semester'].replace(['nan-nan', 'nan-1', 'nan-2', '2024-nan', '2025-nan'], None)
        sched_df['subject_id'] = sched_df['codigo_asignatura'].map(map_subjects)
        sched_df['group_id'] = list(zip(sched_df['grupo'], sched_df['subject_id'], sched_df['semester']))
        sched_df['group_id'] = sched_df['group_id'].map(map_groups)
        
        dia_map = {'LUNES': 1, 'MARTES': 2, 'MIERCOLES': 3, 'JUEVES': 4, 'VIERNES': 5, 'SABADO': 6, 'DOMINGO': 7}
        sched_df['day_of_week'] = sched_df['dia'].astype(str).str.upper().str.strip().map(dia_map)
        sched_df['start_time'] = sched_df['hora_inicial'].astype(str).str.replace('.0','').str.zfill(2) + ':' + sched_df['minutos_hora_inicial'].astype(str).str.replace('.0','').str.zfill(2) + ':00'
        sched_df['end_time'] = sched_df['hora_final'].astype(str).str.replace('.0','').str.zfill(2) + ':' + sched_df['minutos_hora_final'].astype(str).str.replace('.0','').str.zfill(2) + ':00'
        
        sched_df['campus_id'] = sched_df['sede'].map(map_campuses)
        
        sched_final = sched_df[['tenant_id', 'group_id', 'day_of_week', 'start_time', 'end_time', 'aula', 'campus_id']].rename(columns={'aula': 'classroom'}).dropna(subset=['group_id', 'day_of_week'])
        sched_final['day_of_week'] = sched_final['day_of_week'].astype(int)
        sched_final = sched_final.drop_duplicates(subset=['tenant_id', 'group_id', 'day_of_week', 'start_time'])
        
        upsert_batch('group_schedules', sched_final.to_dict('records'), 'tenant_id, group_id, day_of_week, start_time')

    # 8. ACADEMIC PERFORMANCE (NUEVO SISTEMA SEGURO ANTI-PÉRDIDAS)
    if 'definitiva' in df.columns:
        perf_df = df.copy()
        perf_df['semester'] = perf_df['ano'].astype(str).str.replace('.0','') + '-' + perf_df['semestre'].astype(str).str.replace('.0','')
        perf_df['semester'] = perf_df['semester'].replace(['nan-nan', 'nan-1', 'nan-2', '2024-nan', '2025-nan'], None)
        perf_df['subject_id'] = perf_df['codigo_asignatura'].map(map_subjects)
        perf_df['group_id'] = list(zip(perf_df['grupo'], perf_df['subject_id'], perf_df['semester']))
        perf_df['group_id'] = perf_df['group_id'].map(map_groups)
        perf_df['student_id'] = perf_df['documento'].map(map_students)
        perf_df['final_grade'] = pd.to_numeric(perf_df['definitiva'].astype(str).str.replace(',', '.'), errors='coerce')
        
        # Conteo para el radar
        filas_antes = len(perf_df)
        
        # Eliminamos SOLO si el estudiante o el grupo no existen en la BD. (No eliminamos si no hay nota)
        perf_seguro = perf_df.dropna(subset=['student_id', 'group_id']).copy()
        
        filas_despues = len(perf_seguro)
        perdidos = filas_antes - filas_despues
        
        # Alerta en consola si detectamos agujeros
        if perdidos > 0:
            logger.warning(f"⚠️ ALERTA DE DATOS: Se descartaron {perdidos} registros en este lote porque no cruzaron el 'student_id' o 'group_id'.")

        perf_seguro['final_grade'] = perf_seguro['final_grade'].fillna(0.0)
        
        perf_final = perf_seguro[['tenant_id', 'student_id', 'group_id', 'final_grade']].drop_duplicates(['tenant_id', 'student_id', 'group_id'])
        upsert_batch('academic_performance', perf_final.to_dict('records'), 'tenant_id, student_id, group_id')

# --- NUEVO: chunk_size aumentado a 1000 para mayor velocidad ---
def run_etl_pipeline(file_path: str, chunk_size: int = 1000):
    wb = load_workbook(filename=file_path, read_only=True, data_only=True)
    ws = wb.active
    row_iterator = ws.iter_rows(values_only=True)
    headers = clean_column_names(next(row_iterator))
    chunk = []
    for row in row_iterator:
        chunk.append(row)
        if len(chunk) == chunk_size:
            process_and_upload(pd.DataFrame(chunk, columns=headers))
            chunk = []
    if chunk: process_and_upload(pd.DataFrame(chunk, columns=headers))
    logger.info("✅ Pipeline completado exitosamente.")

if __name__ == "__main__":
    archivo_objetivo = "Desarrollo_Curricular_2025_2.xlsx"
    
    # Comprobación de existencia para evitar que el script falle silenciosamente
    if os.path.exists(archivo_objetivo):
        logger.info(f"🚀 Iniciando inyección con protección de datos: {archivo_objetivo}")
        run_etl_pipeline(archivo_objetivo, chunk_size=1000)
    else:
        logger.error(f"❌ No se encontró el archivo '{archivo_objetivo}' en la carpeta actual.")